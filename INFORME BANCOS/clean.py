import pandas as pd
from openpyxl import load_workbook
from path_utils import UPLOAD_FOLDER, NEW_FILES_LIST
import re

patron_final = re.compile(r'^.+ - \d{2}-\d{2}-\d{4}\.xlsx$', re.IGNORECASE)
if NEW_FILES_LIST.exists():
    with open(NEW_FILES_LIST, 'r', encoding='utf-8') as fh:
        nuevos = {l.strip() for l in fh if l.strip()}
else:
    nuevos = None

for f in UPLOAD_FOLDER.glob('*.xlsx'):
    archivo = f.name
    ruta_archivo = str(f)
    if patron_final.match(archivo) and (not nuevos or archivo not in nuevos):
        continue
    try:
        wb = load_workbook(ruta_archivo, data_only=True)
        ws = wb.active
        valor_a1 = str(ws["A1"].value).strip() if ws["A1"].value else ""
        if valor_a1.lower() == "cuenta":
            print(f"'{archivo}' ya tiene 'Cuenta' en A1. Se deja sin modificar.")
        else:
            df = pd.read_excel(ruta_archivo, header=5)
            df.to_excel(ruta_archivo, index=False)
            print(f"Encabezado actualizado en: {archivo}")
    except Exception as e:
        print(f"Error procesando '{archivo}': {e}")
