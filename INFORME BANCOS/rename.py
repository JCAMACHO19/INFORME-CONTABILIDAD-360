import re
import json
from openpyxl import load_workbook
from path_utils import UPLOAD_FOLDER, NEW_FILES_LIST, FECHAS_INICIALES_JSON

def main():
    patron_final = re.compile(r'^.+ - \d{2}-\d{2}-\d{4}\.xlsx$', re.IGNORECASE)
    nuevos = []
    # Cargar fechas iniciales existentes si el archivo ya existe
    if FECHAS_INICIALES_JSON.exists():
        try:
            with open(FECHAS_INICIALES_JSON, 'r', encoding='utf-8') as fh:
                fechas_iniciales = json.load(fh)
        except Exception:
            fechas_iniciales = {}
    else:
        fechas_iniciales = {}
    for f in UPLOAD_FOLDER.glob('*.xlsx'):
        archivo = f.name
        if patron_final.match(archivo):
            print(f"↪️ Ya con nombre final, se omite: {archivo}")
            continue
        ruta_archivo = str(f)
        try:
            wb = load_workbook(ruta_archivo, data_only=True)
            ws = wb.active
            fila_nombre = str(ws["A2"].value)
            if fila_nombre and "NIT" in fila_nombre:
                nombre_parte1 = fila_nombre.split("NIT")[0].strip()
            else:
                nombre_parte1 = "SIN_NOMBRE"
            fila_fecha = str(ws["A4"].value)
            # Extraer Fecha Final
            match_fecha_final = re.search(r"Fecha Final:\s*([0-9]{2}/[0-9]{2}/[0-9]{4})", fila_fecha)
            fecha_final = match_fecha_final.group(1) if match_fecha_final else "FECHA_NO_ENCONTRADA"
            # Extraer Fecha Inicial
            match_fecha_inicial = re.search(r"Fecha Inicial:\s*([0-9]{2}/[0-9]{2}/[0-9]{4})", fila_fecha)
            fecha_inicial = match_fecha_inicial.group(1) if match_fecha_inicial else ""
            fecha_final_safe = fecha_final.replace('/', '-') if fecha_final != 'FECHA_NO_ENCONTRADA' else ''
            nuevo_nombre = f"{nombre_parte1.strip()} - {fecha_final_safe}.xlsx" if fecha_final_safe else f"{nombre_parte1.strip()}.xlsx"
            nueva_ruta = f.with_name(nuevo_nombre)
            if not nueva_ruta.exists():
                f.rename(nueva_ruta)
                nuevos.append(nuevo_nombre)
                if fecha_inicial:
                    fechas_iniciales[nuevo_nombre] = fecha_inicial
                print(f"Renombrado: '{archivo}' → '{nuevo_nombre}' (Fecha Inicial: {fecha_inicial or 'NO_ENCONTRADA'})")
            else:
                # Si ya existe y no tiene fecha inicial registrada, intentar registrar
                if fecha_inicial and nuevo_nombre not in fechas_iniciales:
                    fechas_iniciales[nuevo_nombre] = fecha_inicial
                print(f"Archivo ya existe: '{nuevo_nombre}' → se omite. (Fecha Inicial: {fecha_inicial or 'NO_ENCONTRADA'})")
        except Exception as e:
            print(f"Error procesando '{archivo}': {e}")
    # Guardar lista de nuevos
    try:
        with open(NEW_FILES_LIST, 'w', encoding='utf-8') as fh:
            for n in nuevos:
                fh.write(n + '\n')
    except Exception as e:
        print(f"⚠️ No se pudo escribir lista de nuevos archivos: {e}")
    # Guardar fechas iniciales persistentes
    try:
        with open(FECHAS_INICIALES_JSON, 'w', encoding='utf-8') as fh:
            json.dump(fechas_iniciales, fh, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️ No se pudo escribir fechas iniciales: {e}")

if __name__ == '__main__':
    main()
