import pandas as pd
from openpyxl import load_workbook
from path_utils import UPLOAD_FOLDER, NEW_FILES_LIST
import re

patron_final = re.compile(r'^.+ - \d{2}-\d{2}-\d{4}\.xlsx$', re.IGNORECASE)
if NEW_FILES_LIST.exists():
    with open(NEW_FILES_LIST, 'r', encoding='utf-8') as f:
        nuevos = {l.strip() for l in f if l.strip()}
else:
    nuevos = None  # None = procesar solo los que NO cumplen patrón y aún no están finalizados

for file_path in UPLOAD_FOLDER.glob('*.xlsx'):
    archivo = file_path.name
    # Regla de omisión: si nombre final y no está en "nuevos" (o no hay nuevos definidos), saltar
    if patron_final.match(archivo) and (not nuevos or archivo not in nuevos):
        continue
    ruta_archivo = str(file_path)

    try:
        # Usamos openpyxl para leer el valor exacto de la celda A1
        wb = load_workbook(ruta_archivo, data_only=True)
        ws = wb.active
        valor_a1 = str(ws["A1"].value).strip() if ws["A1"].value else ""
        if valor_a1.lower() == "cuenta contable":
            print(f"'{archivo}' ya tiene 'Cuenta' en A1. Se deja sin modificar.")
        else:
            # Leer Excel a partir de la fila 8 (header=7 en pandas)
            df = pd.read_excel(ruta_archivo, header=7)
            df.to_excel(ruta_archivo, index=False)
            print(f"Encabezado actualizado en: {archivo}")

    except Exception as e:
        print(f"Error procesando '{archivo}': {e}")
