import os
import re
from openpyxl import load_workbook
from path_utils import UPLOAD_FOLDER, NEW_FILES_LIST

def main():
    archivos_excel = [f for f in UPLOAD_FOLDER.glob('*.xlsx')]
    patron_final = re.compile(r'^.+ - \d{2}-\d{2}-\d{4}\.xlsx$', re.IGNORECASE)
    nuevos = []
    for file_path in archivos_excel:
        archivo = file_path.name
        # Idempotencia: si ya cumple el patrón final, se omite
        if patron_final.match(archivo):
            print(f"↪️ Ya con nombre final, se omite: {archivo}")
            continue
        ruta_archivo = str(file_path)
        try:
            wb = load_workbook(ruta_archivo, data_only=True)
            ws = wb.active

            fila_nombre = str(ws["A2"].value)
            if fila_nombre and "NIT" in fila_nombre:
                nombre_parte1 = fila_nombre.split("NIT")[0].strip()
            else:
                nombre_parte1 = "SIN_NOMBRE"

            fila_fecha = str(ws["A4"].value)
            match_fecha = re.search(r"-\s*([0-9]{2}/[0-9]{2}/[0-9]{4})", fila_fecha)
            fecha_final = match_fecha.group(1) if match_fecha else "FECHA_NO_ENCONTRADA"

            fecha_final_safe = fecha_final.replace("/", "-") if fecha_final != "FECHA_NO_ENCONTRADA" else ""
            # Normalizar parte empresa: espacios múltiples y guiones repetidos
            nombre_parte1_norm = re.sub(r"\s+", " ", nombre_parte1).strip()
            nombre_parte1_norm = re.sub(r"-{2,}", "-", nombre_parte1_norm).strip(" -")

            if fecha_final_safe:
                base_nombre = f"{nombre_parte1_norm} - {fecha_final_safe}"
            else:
                # Si no hay fecha, no forzamos renombrado a formato incompleto; usar solo nombre
                base_nombre = nombre_parte1_norm if nombre_parte1_norm else archivo.replace('.xlsx','')

            # Limpiar espacios dobles restantes
            base_nombre = re.sub(r"\s{2,}", " ", base_nombre).strip()
            # Asegurar que la fecha quede exactamente con patrón dd-mm-aaaa sin espacios alrededor de los guiones
            base_nombre = re.sub(r"(\d{2})\s*-\s*(\d{2})\s*-\s*(\d{4})", r"\1-\2-\3", base_nombre)

            nuevo_nombre = f"{base_nombre}.xlsx"
            nueva_ruta = str(UPLOAD_FOLDER / nuevo_nombre)

            if not os.path.exists(nueva_ruta):
                os.rename(ruta_archivo, nueva_ruta)
                nuevos.append(nuevo_nombre)
                print(f"Renombrado: '{archivo}' → '{nuevo_nombre}'")
            else:
                print(f"Archivo ya existe: '{nuevo_nombre}' → se omite.")
        except Exception as e:
            print(f"Error procesando '{archivo}': {e}")
    # Guardar lista de nuevos (puede quedar vacía)
    try:
        with open(NEW_FILES_LIST, 'w', encoding='utf-8') as f:
            for n in nuevos:
                f.write(n + '\n')
    except Exception as e:
        print(f"⚠️ No se pudo escribir lista de nuevos archivos: {e}")

if __name__ == "__main__":
    main()
